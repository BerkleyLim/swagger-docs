import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 🔹 Swagger JSON 파일 로드
with open("api-docs.json", "r", encoding="utf-8") as f:
    swagger_data = json.load(f)

# 🔹 엑셀 워크북 생성
wb = Workbook()

# 🔹 기본 스타일 설정
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# 🔹 API 경로별로 시트 생성
for path, methods in swagger_data["paths"].items():
    for method, details in methods.items():
        api_name = details.get("summary", path)  # API 이름 (없으면 경로 사용)
        sheet_name = f"{method.upper()} {path}"[:31]  # 엑셀 시트 이름은 31자 제한

        ws = wb.create_sheet(title=sheet_name)
        ws.append(["인터페이스 정의서"])
        ws.append([])
        ws.append(["인터페이스 ID", "송신", "수신", "내용"])
        ws.append([sheet_name, "Chub", "Legacy(SCIS)", api_name])
        ws.append([])

        # 🔹 INPUT 섹션 작성
        ws.append(["INPUT"])
        ws.append(["순번", "파라미터 코드", "파라미터 명", "필수 여부", "자료형", "데이터 크기", "샘플 데이터", "비고"])
        input_idx = 1

        # 🔹 요청 파라미터 (parameters) 추출
        if "parameters" in details:
            for param in details["parameters"]:
                ws.append([
                    input_idx, param["name"], param.get("description", "-"),
                    "O" if param.get("required", False) else "",
                    param.get("schema", {}).get("type", "Unknown"), "-", "-", "-"
                ])
                input_idx += 1

        # 🔹 요청 바디 (requestBody) 추출
        if "requestBody" in details:
            request_body = details["requestBody"].get("content", {}).get("application/json", {})
            schema = request_body.get("schema", {})
            if schema.get("type") == "object" and "properties" in schema:
                for key, value in schema["properties"].items():
                    ws.append([
                        input_idx, key, value.get("description", "-"),
                        "O", value.get("type", "Unknown"), "-", "-", "-"
                    ])
                    input_idx += 1

        ws.append([])

        # 🔹 OUTPUT 섹션 작성
        ws.append(["OUTPUT"])
        ws.append(["순번", "파라미터 코드", "파라미터 명", "필수 여부", "자료형", "데이터 크기", "샘플 데이터", "비고"])
        output_idx = 1

        # 🔹 응답 데이터 (responses) 추출
        if "responses" in details:
            for status_code, response in details["responses"].items():
                response_content = response.get("content", {})
                if "application/json" in response_content:
                    schema = response_content["application/json"].get("schema", {})

                    # 🔹 schema가 object일 경우
                    if schema.get("type") == "object" and "properties" in schema:
                        for key, value in schema["properties"].items():
                            ws.append([
                                output_idx, key, value.get("description", "-"),
                                "O", value.get("type", "Unknown"), "-", "-"
                            ])
                            output_idx += 1

                    # 🔹 example 데이터 처리
                    elif "example" in response_content["application/json"]:
                        example_data = response_content["application/json"]["example"]
                        for key, value in example_data.items():
                            ws.append([
                                output_idx, key, "-",
                                "O", type(value).__name__, "-", str(value)
                            ])
                            output_idx += 1

        ws.append([])

        # 🔹 테스트 데이터 작성
        ws.append(["테스트 데이터(1)", "테스트 데이터(2)"])
        test_data_idx = 1
        if "responses" in details:
            for response in details["responses"].values():
                example_data = response.get("content", {}).get("application/json", {}).get("example", {})
                for key, value in example_data.items():
                    ws.append([
                        f'"{key}" : "{value}"'
                    ])
                    test_data_idx += 1

        # 🔹 스타일 적용 (헤더 부분)
        for col in range(1, 10):
            ws.cell(row=7, column=col).font = bold_font
            ws.cell(row=7, column=col).alignment = center_align
            ws.cell(row=7, column=col).fill = gray_fill

# 🔹 기본 시트 삭제 (첫 번째 빈 시트)
wb.remove(wb["Sheet"])

# 🔹 파일 저장
wb.save("interface.xlsx")
print("✅ 엑셀 인터페이스 정의서 (API 단위 시트 분리) 생성 완료! 🚀")